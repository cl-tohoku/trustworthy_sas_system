from pathlib import Path
import sys
import string
import json
import openpyxl

sys.path.append("..")
from library.structure import Script
from library.parser import JapaneseParser
from library.structure import PromptConfig
from library.util import Util


class Toppan:
    def __init__(self):
        pass

    def load_sheet(self, script_path):
        book = openpyxl.load_workbook(script_path)
        active_sheet = book.active
        return active_sheet

    def extract_column(self, sheet, column_id):
        raw_column = [sheet.cell(column=column_id, row=i).value for i in range(3, sheet.max_row)]
        return list(filter(None, raw_column))

    def extract_data(self, sheet):
        texts = self.extract_column(sheet, column_id=2)
        score_a = self.extract_column(sheet, column_id=35)
        score_b = self.extract_column(sheet, column_id=36)
        score_c = self.extract_column(sheet, column_id=37)
        return texts, (score_a, score_b, score_c)

    def symbol_to_score(self, symbol, triangle=False):
        additional_score = 1.0 if triangle else 0.0
        if symbol == "〇":
            return 1.0 + additional_score
        elif symbol == "△":
            return 0.0 + additional_score
        elif symbol == "×":
            return 0.0
        else:
            raise RuntimeError("Invalid Score Symbol")

    def to_scripts(self, texts, score_tuple):
        parser = JapaneseParser()
        scripts = []
        for text, a, b, c in zip(texts, *score_tuple):
            script = Script()
            script.text = parser.mecab(text)
            a = self.symbol_to_score(a)
            b = self.symbol_to_score(b)
            c = self.symbol_to_score(c, triangle=True)
            script.score_vector = [a, b, c]
            script.score = a + b + c
            script.annotation_matrix = None
            scripts.append(script)
        return scripts

    def make_prompt(self):
        prompt = PromptConfig()
        prompt.scoring_item_num = 3
        prompt.deduction_spl = False
        prompt.deduction_eos = False
        prompt.max_scores = [2, 2, 2]
        return prompt

    def __call__(self, config):
        script_path = config.script_path
        sheet = self.load_sheet(script_path)
        texts, score_tuple = self.extract_data(sheet)
        scripts = self.to_scripts(texts, score_tuple)
        prompt = self.make_prompt()
        return scripts, prompt
